#!/usr/bin/env python3
"""
已选课程的Excel导入/导出功能
"""

import pandas as pd
import os
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from .models import Course, TimeSlot, SelectedCourse
from .data_loader import CourseDataLoader


class SelectedCourseExporter:
    """已选课程导出器"""

    def __init__(self, original_course_file: str = None):
        """
        初始化导出器

        Args:
            original_course_file: 原始课程一览表文件路径，用于动态获取列名
        """
        self.max_time_slots = 5  # 预设最多5个时间段列
        self.original_course_file = original_course_file
        self.original_columns = None

    def _load_original_columns(self) -> List[str]:
        """动态加载原始课程一览表的列名"""
        if self.original_columns is not None:
            return self.original_columns

        if self.original_course_file and os.path.exists(self.original_course_file):
            try:
                import pandas as pd

                df = pd.read_excel(self.original_course_file, nrows=0)  # 只读取列名
                self.original_columns = list(df.columns)
                print(f"✓ 从原始文件加载列名: {len(self.original_columns)} 列")
                return self.original_columns
            except Exception as e:
                print(f"⚠️ 无法读取原始文件列名: {e}")

        # 如果无法读取原始文件，使用默认列名
        default_columns = [
            "课程编码",
            "课程名称",
            "开课院系",
            "课程类别",
            "班次",
            "校区",
            "任课教师",
            "学分",
            "学时",
        ]
        print(f"⚠️ 使用默认列名: {len(default_columns)} 列")
        self.original_columns = default_columns
        return self.original_columns

    def export_to_excel(
        self, selected_courses: List[SelectedCourse], file_path: str
    ) -> bool:
        """导出已选课程到Excel文件（三个工作表格式）"""
        try:
            print(f"正在导出已选课程到: {file_path}")

            if not selected_courses:
                print("❌ 没有已选课程可以导出")
                return False

            # 创建Excel工作簿
            from openpyxl import Workbook

            wb = Workbook()

            # 删除默认工作表
            wb.remove(wb.active)

            # 创建三个工作表
            self._create_course_list_sheet(wb, selected_courses)
            self._create_weekly_schedule_sheet(wb, selected_courses)
            self._create_statistics_sheet(wb, selected_courses)

            # 保存文件
            wb.save(file_path)

            print(
                f"✅ 成功导出 {len(selected_courses)} 门课程到 {Path(file_path).name}"
            )
            print("   包含：课程列表、周课表、统计信息 三个工作表")
            return True

        except Exception as e:
            print(f"❌ 导出失败: {e}")
            import traceback

            traceback.print_exc()
            return False

    def export_with_credit_statistics(
        self,
        selected_courses: List[SelectedCourse],
        file_path: str,
        credit_manager=None,
    ) -> bool:
        """导出已选课程到Excel文件，包含学分统计信息"""
        try:
            print(f"正在导出已选课程（含学分统计）到: {file_path}")

            if not selected_courses:
                print("❌ 没有已选课程可以导出")
                return False

            # 创建Excel工作簿
            from openpyxl import Workbook

            wb = Workbook()

            # 导出课程列表
            self._export_course_list_to_workbook(wb, selected_courses)

            # 导出学分统计（如果提供了credit_manager）
            if credit_manager:
                self._export_credit_statistics_to_workbook(
                    wb, selected_courses, credit_manager
                )

            # 保存文件
            wb.save(file_path)
            print(
                f"✅ 成功导出 {len(selected_courses)} 门课程到 {Path(file_path).name}"
            )
            return True

        except Exception as e:
            print(f"❌ 导出失败: {e}")
            return False

    def _get_excel_columns(self) -> List[str]:
        """获取Excel列名（动态复制原始课程一览表列名，并添加排课相关信息）"""
        # 动态获取原始课程一览表的列名
        base_columns = self._load_original_columns()

        # 排课相关的额外信息
        scheduling_columns = ["自定义类别", "是否线上"]

        # 添加拆分的时间段列
        time_columns = []
        for i in range(self.max_time_slots):
            time_columns.extend([f"星期几{i + 1}", f"节次{i + 1}", f"周次{i + 1}"])

        return base_columns + scheduling_columns + time_columns

    def _convert_selected_course_to_row(
        self, selected_course: SelectedCourse
    ) -> List[Any]:
        """将SelectedCourse对象转换为Excel行数据（动态适配原始课程一览表格式）"""
        course = selected_course.course
        original_columns = self._load_original_columns()

        # 动态构建基础信息行：按照原始课程一览表的列顺序
        row = []
        for col_name in original_columns:
            # 根据列名动态映射数据
            if col_name in ["课程编码", "课程代码"]:
                row.append(course.code)
            elif col_name in ["课程名称", "课程名"]:
                row.append(course.name)
            elif col_name in ["开课院系", "院系", "开课单位"]:
                row.append(course.department)
            elif col_name in ["课程类别", "类别", "课程性质"]:
                # 使用自定义类别而不是原始类别，确保与CreditManager一致
                row.append(selected_course.custom_category)
            elif col_name in ["班次", "班级", "班号"]:
                row.append(selected_course.class_num)
            elif col_name in ["校区", "上课地点"]:
                row.append(course.campus)
            elif col_name in ["任课教师", "教师", "主讲教师"]:
                row.append(course.teacher)
            elif col_name in ["学分", "学分数"]:
                row.append(course.credits)
            elif col_name in ["学时", "总学时", "课时"]:
                row.append(course.hours)
            else:
                # 对于其他列（如选课说明等），尝试从course对象获取，如果没有则为空
                value = getattr(course, col_name.lower().replace(" ", "_"), "")
                row.append(value)

        # 排课相关的额外信息
        row.extend(
            [
                selected_course.custom_category,  # 自定义类别
                "是" if selected_course.is_online else "否",  # 是否线上
            ]
        )

        # 时间段信息 - 拆分为独立列
        for i in range(self.max_time_slots):
            if i < len(selected_course.time_slots):
                time_slot = selected_course.time_slots[i]
                # 星期几（1-7）
                row.append(time_slot.weekday)
                # 节次（格式：开始-结束）
                row.append(f"{time_slot.start_section}-{time_slot.end_section}")
                # 周次（格式化为字符串）
                row.append(self._format_weeks_for_excel(time_slot.weeks))
            else:
                # 没有时间段时，三列都显示空白
                row.extend(["", "", ""])

        return row

    def _create_course_list_sheet(self, wb, selected_courses):
        """创建课程列表工作表 - 直接复制已选课程表格式"""
        ws = wb.create_sheet("课程列表")

        # 使用与已选课程.xlsx完全相同的列标题
        headers = [
            "课程编码",
            "课程名称",
            "开课院系",
            "课程类别",
            "班次",
            "校区",
            "任课教师",
            "学分",
            "学时",
            "自定义类别",
            "是否线上",
            "星期几1",
            "节次1",
            "周次1",
            "星期几2",
            "节次2",
            "周次2",
            "星期几3",
            "节次3",
            "周次3",
            "星期几4",
            "节次4",
            "周次4",
            "星期几5",
            "节次5",
            "周次5",
        ]

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # 写入数据 - 完全按照已选课程表的格式
        for row_idx, selected_course in enumerate(selected_courses, 2):
            course = selected_course.course

            # 基本信息
            ws.cell(row=row_idx, column=1, value=course.code)  # 课程编码
            ws.cell(row=row_idx, column=2, value=course.name)  # 课程名称
            ws.cell(row=row_idx, column=3, value=course.department)  # 开课院系
            ws.cell(row=row_idx, column=4, value=course.category)  # 课程类别
            ws.cell(row=row_idx, column=5, value=selected_course.class_num)  # 班次
            ws.cell(row=row_idx, column=6, value=course.campus)  # 校区
            ws.cell(row=row_idx, column=7, value=course.teacher)  # 任课教师
            ws.cell(row=row_idx, column=8, value=course.credits)  # 学分
            ws.cell(row=row_idx, column=9, value=course.hours)  # 学时
            ws.cell(
                row=row_idx, column=10, value=selected_course.custom_category
            )  # 自定义类别
            ws.cell(
                row=row_idx,
                column=11,
                value="是" if selected_course.is_online else "否",
            )  # 是否线上

            # 时间信息 - 最多5个时间段
            time_slots = selected_course.time_slots[:5]  # 最多取5个时间段

            for i, time_slot in enumerate(time_slots):
                col_base = 12 + i * 3  # 从第12列开始，每个时间段占3列
                ws.cell(
                    row=row_idx, column=col_base, value=float(time_slot.weekday)
                )  # 星期几
                ws.cell(
                    row=row_idx,
                    column=col_base + 1,
                    value=f"{time_slot.start_section}-{time_slot.end_section}",
                )  # 节次
                ws.cell(
                    row=row_idx,
                    column=col_base + 2,
                    value=f"{min(time_slot.weeks)}-{max(time_slot.weeks)}",
                )  # 周次

            # 如果时间段不足5个，其余列留空（Excel会自动处理为空值）

    def _create_weekly_schedule_sheet(self, wb, selected_courses):
        """创建周课表工作表 - 修复节次时间对应关系"""
        ws = wb.create_sheet("周课表")

        # 定义时间节次对应关系 - 修复：每个节次单独处理
        time_slots = [
            ("第1-2节\n08:00-09:40", 1, 2),
            ("第3-4节\n10:00-11:40", 3, 4),
            ("第5-6节\n13:00-14:40", 5, 6),
            ("第7-8节\n15:00-16:40", 7, 8),
            ("第9-10节\n18:00-19:40", 9, 10),
        ]

        # 定义列标题
        headers = ["节次", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]

        # 写入表头
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # 创建课表数据结构
        schedule = {}
        for time_name, start_section, end_section in time_slots:
            schedule[time_name] = [""] * 7  # 7天

        # 填充课程数据 - 修复：正确处理节次范围
        for selected_course in selected_courses:
            course = selected_course.course
            for time_slot in selected_course.time_slots:
                weekday = time_slot.weekday - 1  # 转换为0-6
                if 0 <= weekday <= 6:
                    # 找到对应的时间段 - 修复：检查节次是否在范围内
                    for time_name, start_section, end_section in time_slots:
                        # 检查课程的节次是否与当前时间段有重叠
                        course_start = time_slot.start_section
                        course_end = time_slot.end_section

                        # 判断是否有重叠：课程节次范围与时间段范围有交集
                        if not (
                            course_end < start_section or course_start > end_section
                        ):
                            # 构建课程信息
                            weeks_str = (
                                f"{min(time_slot.weeks)}-{max(time_slot.weeks)}"
                                if time_slot.weeks
                                else ""
                            )
                            course_info = f"{course.name}\n{course.teacher}\n{course.campus}\n第{course_start}-{course_end}节\n{weeks_str}周"

                            # 如果该时间段已有课程，合并显示
                            if schedule[time_name][weekday]:
                                schedule[time_name][weekday] += f"\n\n{course_info}"
                            else:
                                schedule[time_name][weekday] = course_info

        # 写入课表数据
        for row_idx, (time_name, courses) in enumerate(schedule.items(), 2):
            ws.cell(row=row_idx, column=1, value=time_name)  # 节次
            for col_idx, course_info in enumerate(courses, 2):
                ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=course_info if course_info else "",
                )

    def _create_statistics_sheet(self, wb, selected_courses):
        """创建统计信息工作表"""
        ws = wb.create_sheet("统计信息")

        # 写入表头
        ws.cell(row=1, column=1, value="统计项")
        ws.cell(row=1, column=2, value="数值")

        # 计算统计数据
        total_courses = len(selected_courses)
        total_credits = sum(sc.course.credits for sc in selected_courses)
        total_hours = sum(sc.course.hours for sc in selected_courses)

        # 按类别统计
        from collections import defaultdict

        category_stats = defaultdict(lambda: {"count": 0, "credits": 0})
        campus_stats = defaultdict(int)

        for sc in selected_courses:
            category = sc.custom_category
            category_stats[category]["count"] += 1
            category_stats[category]["credits"] += sc.course.credits
            campus_stats[sc.course.campus] += 1

        # 写入统计数据
        row_idx = 2

        # 基本统计
        ws.cell(row=row_idx, column=1, value="总课程数")
        ws.cell(row=row_idx, column=2, value=total_courses)
        row_idx += 1

        ws.cell(row=row_idx, column=1, value="总学分")
        ws.cell(row=row_idx, column=2, value=total_credits)
        row_idx += 1

        ws.cell(row=row_idx, column=1, value="总学时")
        ws.cell(row=row_idx, column=2, value=total_hours)
        row_idx += 1

        ws.cell(row=row_idx, column=1, value="平均学分")
        ws.cell(
            row=row_idx,
            column=2,
            value=round(total_credits / total_courses, 2) if total_courses > 0 else 0,
        )
        row_idx += 1

        # 空行
        row_idx += 1

        # 类别统计
        ws.cell(row=row_idx, column=1, value="=== 按类别统计 ===")
        row_idx += 1

        for category, stats in category_stats.items():
            ws.cell(row=row_idx, column=1, value=f"{category}课程数")
            ws.cell(row=row_idx, column=2, value=stats["count"])
            row_idx += 1

            ws.cell(row=row_idx, column=1, value=f"{category}学分")
            ws.cell(row=row_idx, column=2, value=stats["credits"])
            row_idx += 1

        # 空行
        row_idx += 1

        # 校区统计
        ws.cell(row=row_idx, column=1, value="=== 按校区统计 ===")
        row_idx += 1

        for campus, count in campus_stats.items():
            ws.cell(row=row_idx, column=1, value=f"{campus}课程数")
            ws.cell(row=row_idx, column=2, value=count)
            row_idx += 1

    def _get_course_time_info(self, selected_course):
        """获取课程时间信息"""
        if not selected_course.time_slots:
            return {"weeks": "", "weekday": "", "sections": ""}

        # 取第一个时间段的信息
        time_slot = selected_course.time_slots[0]

        # 周次
        weeks = (
            f"{min(time_slot.weeks)}-{max(time_slot.weeks)}" if time_slot.weeks else ""
        )

        # 星期
        weekdays = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]
        weekday = weekdays[time_slot.weekday] if 1 <= time_slot.weekday <= 7 else ""

        # 节次
        sections = f"{time_slot.start_section}-{time_slot.end_section}"

        return {"weeks": weeks, "weekday": weekday, "sections": sections}

    def _export_course_list_to_workbook(self, wb, selected_courses):
        """导出课程列表到工作簿"""
        ws = wb.active
        ws.title = "课程列表"

        # 准备数据
        data = []
        for selected_course in selected_courses:
            row = self._convert_selected_course_to_row(selected_course)
            data.append(row)

        # 获取列名
        columns = self._get_excel_columns()

        # 写入表头
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    def _export_credit_statistics_to_workbook(
        self, wb, selected_courses, credit_manager
    ):
        """导出学分统计到工作簿"""
        ws = wb.create_sheet("学分统计")

        # 统计各类别的学分
        from collections import defaultdict

        category_stats = defaultdict(float)
        for sc in selected_courses:
            category_stats[sc.custom_category] += sc.course.credits

        # 写入表头
        headers = [
            "课程类别",
            "要求学分",
            "已修学分",
            "新选学分",
            "总计学分",
            "完成状态",
            "超出学分",
        ]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        # 写入统计数据
        row_idx = 2
        total_required = 0
        total_completed = 0
        total_selected = 0
        total_all = 0

        for category, requirement in credit_manager.requirements.items():
            selected_credits = category_stats.get(category, 0)
            completed_credits = requirement.completed_credits
            required_credits = requirement.required_credits
            total_credits = completed_credits + selected_credits
            excess_credits = max(0, total_credits - required_credits)

            # 完成状态
            if total_credits >= required_credits:
                status = "✅ 已完成"
            else:
                status = f"❌ 缺少 {required_credits - total_credits} 学分"

            # 写入数据
            ws.cell(row=row_idx, column=1, value=category)
            ws.cell(row=row_idx, column=2, value=required_credits)
            ws.cell(row=row_idx, column=3, value=completed_credits)
            ws.cell(row=row_idx, column=4, value=selected_credits)
            ws.cell(row=row_idx, column=5, value=total_credits)
            ws.cell(row=row_idx, column=6, value=status)
            ws.cell(row=row_idx, column=7, value=excess_credits)

            # 累计统计
            total_required += required_credits
            total_completed += completed_credits
            total_selected += selected_credits
            total_all += total_credits

            row_idx += 1

        # 添加总计行
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="总计")
        ws.cell(row=row_idx, column=2, value=total_required)
        ws.cell(row=row_idx, column=3, value=total_completed)
        ws.cell(row=row_idx, column=4, value=total_selected)
        ws.cell(row=row_idx, column=5, value=total_all)

        # 添加说明
        row_idx += 2
        ws.cell(row=row_idx, column=1, value="说明：")
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="• 要求学分：培养方案规定的最低学分要求")
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="• 已修学分：用户设置的已完成学分")
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="• 新选学分：本次排课选择的学分")
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="• 总计学分：已修学分 + 新选学分")
        row_idx += 1
        ws.cell(
            row=row_idx,
            column=1,
            value="• 超出学分：超过要求学分的部分（允许适度超出）",
        )

    def _format_weeks_for_excel(self, weeks: List[int]) -> str:
        """将周次列表格式化为Excel友好的字符串"""
        if not weeks:
            return ""

        # 找连续的周次
        ranges = []
        start = weeks[0]
        end = start

        for i in range(1, len(weeks)):
            if weeks[i] == end + 1:
                end = weeks[i]
            else:
                if start == end:
                    ranges.append(str(start))
                else:
                    ranges.append(f"{start}-{end}")
                start = end = weeks[i]

        # 添加最后一个范围
        if start == end:
            ranges.append(str(start))
        else:
            ranges.append(f"{start}-{end}")

        return ",".join(ranges)


class SelectedCourseImporter:
    """已选课程导入器"""

    def __init__(self):
        self.max_time_slots = 5
        self.import_report = {}

    def import_from_excel(
        self, file_path: str, course_loader: CourseDataLoader
    ) -> Tuple[List[SelectedCourse], Dict[str, Any]]:
        """从Excel文件导入已选课程"""
        try:
            print(f"正在从Excel导入已选课程: {file_path}")

            # 检查文件是否存在
            if not Path(file_path).exists():
                print(f"❌ 文件不存在: {file_path}")
                return [], {"success": False, "error": "文件不存在"}

            # 检查是否已加载课程一览表
            if not course_loader or not course_loader.get_courses():
                print("❌ 请先加载课程一览表")
                return [], {"success": False, "error": "请先加载课程一览表"}

            # 读取Excel文件
            df = pd.read_excel(file_path)
            print(f"✓ 成功读取Excel文件，共 {len(df)} 条记录")

            # 验证列名
            if not self._validate_columns(df):
                return [], {"success": False, "error": "Excel文件格式不正确"}

            # 建立课程索引
            course_index = self._build_course_index(course_loader.get_courses())

            # 转换数据
            selected_courses = []
            failed_count = 0

            for index, row in df.iterrows():
                try:
                    selected_course = self._convert_row_to_selected_course(
                        row, course_index
                    )
                    if selected_course:
                        selected_courses.append(selected_course)
                except Exception as e:
                    failed_count += 1
                    print(f"⚠️ 第 {index + 1} 行数据转换失败: {e}")

            # 生成导入报告
            self._generate_import_report(len(df), len(selected_courses), failed_count)

            print(f"✅ 导入完成: 成功导入 {len(selected_courses)} 门课程")
            if failed_count > 0:
                print(f"⚠️ {failed_count} 条记录导入失败")

            return selected_courses, self.import_report

        except Exception as e:
            print(f"❌ 导入失败: {e}")
            return [], {"success": False, "error": str(e)}

    def _validate_columns(self, df: pd.DataFrame) -> bool:
        """验证Excel列名（动态验证）"""
        columns = list(df.columns)

        # 检查必需的基础列（至少要有课程编码和班次用于匹配）
        essential_columns = []

        # 课程编码列（多种可能的名称）
        code_columns = ["课程编码", "课程代码", "课程号"]
        code_found = any(col in columns for col in code_columns)
        if not code_found:
            essential_columns.extend(code_columns)

        # 班次列
        class_columns = ["班次", "班级", "班号"]
        class_found = any(col in columns for col in class_columns)
        if not class_found:
            essential_columns.extend(class_columns)

        if essential_columns:
            print(f"❌ Excel文件缺少必需列，需要以下列之一: {essential_columns}")
            return False

        # 检查排课相关列
        required_scheduling_columns = ["自定义类别", "是否线上"]
        missing_scheduling_columns = [
            col for col in required_scheduling_columns if col not in columns
        ]
        if missing_scheduling_columns:
            print(f"❌ Excel文件缺少排课相关列: {missing_scheduling_columns}")
            return False

        # 检查时间段列（至少要有第一组时间段列）
        required_time_columns = ["星期几1", "节次1", "周次1"]
        missing_time_columns = [
            col for col in required_time_columns if col not in columns
        ]
        if missing_time_columns:
            print(f"❌ Excel文件缺少时间段列: {missing_time_columns}")
            print("提示：Excel文件应包含至少一组时间段列（星期几1、节次1、周次1）")
            return False

        print("✓ Excel文件列验证通过")
        return True

    def _build_course_index(self, courses: List[Course]) -> Dict[str, List[Course]]:
        """建立课程索引"""
        course_index = {}
        for course in courses:
            key = f"{course.code}_{course.class_num}"
            if key not in course_index:
                course_index[key] = []
            course_index[key].append(course)
        return course_index

    def _convert_row_to_selected_course(
        self, row, course_index: Dict[str, List[Course]]
    ) -> Optional[SelectedCourse]:
        """将Excel行数据转换为SelectedCourse对象（动态适配列名）"""
        try:
            # 动态查找课程编码列
            course_code = None
            for col_name in ["课程编码", "课程代码", "课程号"]:
                if col_name in row and pd.notna(row[col_name]):
                    course_code = str(row[col_name]).strip()
                    break

            if not course_code:
                print("⚠️ 未找到课程编码")
                return None

            # 动态查找班次列
            class_num = None
            for col_name in ["班次", "班级", "班号"]:
                if col_name in row and pd.notna(row[col_name]):
                    class_num = int(row[col_name])
                    break

            if class_num is None:
                print("⚠️ 未找到班次信息")
                return None

            # 查找对应的课程
            key = f"{course_code}_{class_num}"
            if key not in course_index:
                print(f"⚠️ 未找到课程: {course_code} 班次{class_num}")
                return None

            course = course_index[key][0]  # 取第一个匹配的课程

            # 读取排课相关信息
            is_online = str(row["是否线上"]).strip() == "是"
            is_imported = True  # 从Excel导入的课程都标记为导入课程

            # 处理自定义类别，避免NaN问题
            custom_category_raw = row["自定义类别"]
            if pd.isna(custom_category_raw) or str(
                custom_category_raw
            ).strip().lower() in ["nan", ""]:
                custom_category = (
                    ""  # 空字符串，让SelectedCourse自动分配（可选类型会分配为"nan"）
                )
            else:
                custom_category = str(custom_category_raw).strip()

            # 解析时间段
            time_slots = self._parse_time_slots(row)

            selected_course = SelectedCourse(
                course=course,
                class_num=class_num,
                time_slots=time_slots,
                is_online=is_online,
                custom_category=custom_category,
                is_imported=is_imported,
            )

            return selected_course

        except Exception as e:
            print(f"转换行数据失败: {e}")
            return None

    def _parse_time_slots(self, row) -> List[TimeSlot]:
        """解析时间段信息"""
        time_slots = []

        for i in range(self.max_time_slots):
            weekday_col = f"星期几{i + 1}"
            section_col = f"节次{i + 1}"
            weeks_col = f"周次{i + 1}"

            # 检查是否存在这组时间段列
            if not all(col in row for col in [weekday_col, section_col, weeks_col]):
                break

            # 检查是否有有效数据
            if (
                pd.notna(row[weekday_col])
                and str(row[weekday_col]).strip()
                and pd.notna(row[section_col])
                and str(row[section_col]).strip()
            ):
                try:
                    time_slot = self._parse_time_slot_from_columns(
                        row[weekday_col], row[section_col], row[weeks_col], i + 1
                    )
                    if time_slot:
                        time_slots.append(time_slot)
                except Exception as e:
                    print(f"⚠️ 解析时间段{i + 1}失败: {e}")

        return time_slots

    def _parse_time_slot_from_columns(
        self, weekday_val, section_val, weeks_val, slot_num: int
    ) -> Optional[TimeSlot]:
        """从拆分的列中解析时间段"""
        try:
            # 解析星期几
            weekday = int(weekday_val)
            if not (1 <= weekday <= 7):
                print(
                    f"⚠️ 时间段{slot_num}：星期几必须是1-7之间的数字（当前值：{weekday_val}）"
                )
                return None

            # 解析节次
            section_str = str(section_val).strip()
            if "-" not in section_str:
                print(
                    f"⚠️ 时间段{slot_num}：节次格式应为'开始-结束'（如：1-2）（当前值：{section_val}）"
                )
                return None

            try:
                start_section, end_section = section_str.split("-")
                start_section = int(start_section.strip())
                end_section = int(end_section.strip())

                if not (1 <= start_section <= 10 and 1 <= end_section <= 10):
                    print(
                        f"⚠️ 时间段{slot_num}：节次必须在1-10范围内（当前值：{section_val}）"
                    )
                    return None

                if start_section > end_section:
                    print(
                        f"⚠️ 时间段{slot_num}：开始节次不能大于结束节次（当前值：{section_val}）"
                    )
                    return None

            except ValueError:
                print(
                    f"⚠️ 时间段{slot_num}：节次格式错误，应为数字（当前值：{section_val}）"
                )
                return None

            # 解析周次
            weeks = []
            if pd.notna(weeks_val) and str(weeks_val).strip():
                weeks = self._parse_weeks_string_improved(
                    str(weeks_val).strip(), slot_num
                )
                if not weeks:
                    print(f"⚠️ 时间段{slot_num}：周次解析失败（当前值：{weeks_val}）")
                    return None
            else:
                # 如果周次为空，默认为全学期
                weeks = list(range(1, 21))

            return TimeSlot(
                weekday=weekday,
                start_section=start_section,
                end_section=end_section,
                weeks=weeks,
            )

        except Exception as e:
            print(f"⚠️ 时间段{slot_num}解析失败: {e}")
            return None

    def _parse_weeks_string_improved(self, weeks_str: str, slot_num: int) -> List[int]:
        """改进的周次字符串解析，提供详细错误信息"""
        try:
            weeks = []

            # 按逗号分割
            parts = weeks_str.split(",")

            for part in parts:
                part = part.strip()
                if not part:
                    continue

                if "-" in part:
                    # 连续周次，如"1-10"
                    try:
                        start, end = part.split("-")
                        start = int(start.strip())
                        end = int(end.strip())

                        if not (1 <= start <= 20 and 1 <= end <= 20):
                            print(
                                f"⚠️ 时间段{slot_num}：周次范围必须在1-20之间（当前值：{part}）"
                            )
                            return []

                        if start > end:
                            print(
                                f"⚠️ 时间段{slot_num}：开始周次不能大于结束周次（当前值：{part}）"
                            )
                            return []

                        weeks.extend(range(start, end + 1))
                    except ValueError:
                        print(f"⚠️ 时间段{slot_num}：周次范围格式错误（当前值：{part}）")
                        return []
                else:
                    # 单个周次
                    try:
                        week = int(part)
                        if not (1 <= week <= 20):
                            print(
                                f"⚠️ 时间段{slot_num}：周次必须在1-20范围内（当前值：{part}）"
                            )
                            return []
                        weeks.append(week)
                    except ValueError:
                        print(f"⚠️ 时间段{slot_num}：周次必须是数字（当前值：{part}）")
                        return []

            return sorted(list(set(weeks)))

        except Exception as e:
            print(f"⚠️ 时间段{slot_num}周次解析失败: {weeks_str}, 错误: {e}")
            return []

    def _parse_time_slot_string(self, time_str: str) -> Optional[TimeSlot]:
        """解析时间段字符串，如"周一1-2节(1-10周)" """
        try:
            # 解析格式：周一1-2节(1-10周)
            import re

            # 匹配模式
            pattern = r"周([一二三四五六日])\s*(\d+)-(\d+)节\s*\(([^)]+)\)"
            match = re.match(pattern, time_str)

            if not match:
                print(f"⚠️ 无法解析时间段格式: {time_str}")
                return None

            weekday_name, start_section, end_section, weeks_str = match.groups()

            # 转换星期几
            weekday_map = {
                "一": 1,
                "二": 2,
                "三": 3,
                "四": 4,
                "五": 5,
                "六": 6,
                "日": 7,
            }
            weekday = weekday_map[weekday_name]

            # 解析周次
            weeks = self._parse_weeks_string(weeks_str)

            return TimeSlot(
                weekday=weekday,
                start_section=int(start_section),
                end_section=int(end_section),
                weeks=weeks,
            )

        except Exception as e:
            print(f"解析时间段失败: {time_str}, 错误: {e}")
            return None

    def _parse_weeks_string(self, weeks_str: str) -> List[int]:
        """解析周次字符串，如"1-10周"或"1,3,5周" """
        try:
            weeks = []
            # 移除"周"字
            weeks_str = weeks_str.replace("周", "")

            # 按逗号分割
            parts = weeks_str.split(",")

            for part in parts:
                part = part.strip()
                if "-" in part:
                    # 连续周次，如"1-10"
                    start, end = part.split("-")
                    weeks.extend(range(int(start), int(end) + 1))
                else:
                    # 单个周次
                    weeks.append(int(part))

            return sorted(list(set(weeks)))

        except Exception as e:
            print(f"解析周次失败: {weeks_str}, 错误: {e}")
            return []

    def _generate_import_report(
        self, total_records: int, successful_records: int, failed_records: int
    ):
        """生成导入报告"""
        self.import_report = {
            "success": True,
            "total_records": total_records,
            "successful_records": successful_records,
            "failed_records": failed_records,
            "success_rate": f"{(successful_records / total_records * 100):.1f}%"
            if total_records > 0
            else "0%",
        }

    def get_import_report(self) -> Dict[str, Any]:
        """获取导入报告"""
        return self.import_report
